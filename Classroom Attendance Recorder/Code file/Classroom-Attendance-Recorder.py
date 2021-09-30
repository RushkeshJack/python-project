
# All the required libraries are installed and imported first
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook
from time import sleep
import datetime
from tkinter import *
from tkinter import messagebox
import os

# The Tkinter code of GUI is started here
window = Tk()
window.title("Classroom Attendance Recorder")
window_width = 500
window_height = 500
window.geometry(f'{window_width}x{window_height}+{512}+{192}')
window.geometry("500x500")
window.iconbitmap(r"C:\Classroom-Attendance-Recorder\ICON\attendance.ico")


# 3 input fields are created i.e email, password and input field for Google meet link
Email = Label(window, text="Email : ", font=("arial", 15))
E = Entry(window, font=("arial", 15))
#E.grid(row=1, column=2)
E.place(x=130, y=100)
E.place(width=350, height=35)
Email.place(x=46, y=100)

Password = Label(window, text="Password : ", font=("arial", 15))
P = Entry(window, font=("arial", 15))
P.config(show="*")
#P.grid(row=3, column=2)
P.place(x=130, y=200)
P.place(width=350, height=35)
Password.place(x=5, y=200)

Meet_link = Label(window, text="Meet link : ", font=("arial", 15))
M = Entry(window, font=("arial", 15))
#M.grid(row=5, column=2)
M.place(x=130, y=300)
M.place(width=350, height=35)
Meet_link.place(x=13, y=300)


def install(package):
    os.system("pip install " + str(package))
    #print("Installed", package.upper())

def user_info():  # This function will get executed when 'About Us' Button get clicked

    messagebox.showinfo("About Us", "We are B.tech students and working on a problem i.e. how to make human lives simpler!")


def restore():  # This function is get executed when 'Restore' button get pressed

    workbook = load_workbook('Login and meet link.xlsx')
    sheet = workbook['Login']

    gmail_id = sheet['A2'].value
    gmail_password = sheet['B2'].value
    meet_link = sheet['C2'].value

    E.insert(0, gmail_id)
    P.insert(0, gmail_password)
    P.config(show="*")
    M.insert(0, meet_link)


def submit():  # This function is get executed when 'Submit' button get pressed
    install("openpyxl")
    install("selenium")
    gmail_id = E.get()
    gmail_password = P.get()
    meet_link = M.get()

    # Some necessary things for automation with google driver
    opt = Options()
    opt.add_argument("--disable-infobars")
    opt.add_argument("start-maximized")
    opt.add_experimental_option("excludeSwitches", ['enable-automation'])
    opt.add_experimental_option('excludeSwitches', ['enable-logging'])

    # This part allows the notifications, mic and camera permissions pass the argument 1 to allow and 2 to block
    opt.add_experimental_option("prefs", {
        "profile.default_content_setting_values.media_stream_mic": 1,
        "profile.default_content_setting_values.media_stream_camera": 1,
        "profile.default_content_setting_values.geolocation": 1,
        "profile.default_content_setting_values.notifications": 1
      })

    # Sign in to google
    driver = webdriver.Chrome(options=opt, executable_path=r"C:\Classroom-Attendance-Recorder\ChromeDriver\chromedriver")
    driver.get('https://apps.google.com/intl/en-GB/meet/')
    driver.find_element_by_xpath('/html/body/header/div[1]/div/div[3]/div[1]/div/span[1]/a').click()
    sleep(3)
    driver.find_element_by_xpath('//input[@type="email"]').send_keys(gmail_id)  # entering the gmail id
    driver.find_element_by_xpath('//*[@id="identifierNext"]').click()
    sleep(3)
    driver.find_element_by_xpath('//input[@type="password"]').send_keys(gmail_password)  # entering the password
    driver.find_element_by_xpath('//*[@id="passwordNext"]').click()
    sleep(3)
    driver.find_element_by_css_selector('#i3').send_keys(meet_link)  # Enter a code or link
    sleep(3)
    driver.find_element_by_css_selector('button.VfPpkd-LgbsSe.VfPpkd-LgbsSe-OWXEXe-dgl2Hf.ksBjEc.lKxP2d.cjtUbb').click()  # join
    sleep(5)
    driver.find_element_by_xpath('//*[@id="yDmH0d"]/c-wiz/div/div/div[9]/div[3]/div/div/div[4]/div/div/div[1]/div[1]/div/div[4]/div[1]/div/div/div').click()  # Mic turn off
    sleep(4)
    driver.find_element_by_xpath('//*[@id="yDmH0d"]/c-wiz/div/div/div[9]/div[3]/div/div/div[4]/div/div/div[1]/div[1]/div/div[4]/div[2]/div/div').click()  # Camera turn off
    sleep(4)
    driver.find_element_by_css_selector('div.uArJ5e.UQuaGc.Y5sE8d.uyXBBb.xKiqt').click()  # join now
    sleep(4)
    driver.find_element_by_xpath('//*[@id="ow3"]/div[1]/div/div[9]/div[3]/div[10]/div[3]/div[2]/div/div/div[2]/span/button/i[1]').click()  # participant list
    sleep(2)
    names = driver.find_elements_by_css_selector('span.ZjFb7c')  # participants


    # Next code upto workbook.save() function. This code used for taking attendance i.e.marking present or absent in excel sheet.
    workbook = load_workbook('Google_Attendance.xlsx')
    sheet1 = workbook['Attendance Sheet']
    count_student = 0
    x = datetime.datetime.now()
    date = x.strftime("%x")
    time = x.strftime("%I:%M")
    meridiem = x.strftime("%p")
    str1 = date
    str2 = time + " " + meridiem
    Maxcolumn = sheet1.max_column

    sheet1.cell(row=6, column=Maxcolumn + 1).value = str1
    sheet1.cell(row=7, column=Maxcolumn + 1).value = str2

    # In this code, the names from google meet participant list get saved in 'names' array and these all names one by one compared
    # to names present in Google_Attendace.xlsx excel sheet.
    for name in names:
        for box in range(8, sheet1.max_row + 1):
            if name.text.lower() == (sheet1.cell(row=box, column=1).value.lower()):  # checks if the participant is present in the sheet
                sheet1.cell(row=box, column=Maxcolumn + 1).value = "Present"  # mark present of that participant
                count_student += 1

    for box in range(8, sheet1.max_row + 1):
        if sheet1.cell(row=box, column=Maxcolumn + 1).value == None:
            sheet1.cell(row=box, column=Maxcolumn + 1).value = 'Absent'

    workbook.save('Google_Attendance.xlsx')
    #print('Attendance taken successfully!')


    driver.find_element_by_xpath('//*[@id="ow3"]/div[1]/div/div[9]/div[3]/div[10]/div[2]/div/div[7]/span/button/i').click()
    sleep(4)
    driver.close()
    driver.quit()


menubar = Menu(window)
menubar.add_command(label="About Us", command=user_info)
menubar.add_command(label="Quit", command=window.quit)
window.config(menu=menubar)


submit = Button(window, text="Submit", command=submit, padx=10, pady=5, font=("arial", 15)).place(x=240, y=375)
restore = Button(window, text="Restore", command=restore, padx=5, pady=5, font=("arial", 10), fg="white", bg="light blue").place(x=425, y=10)


window.mainloop()
