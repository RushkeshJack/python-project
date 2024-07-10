import openpyxl
from openpyxl import Workbook

def create_workbook():
    wb = Workbook()
    ws = wb.active
    ws.append(['Student_ID', 'Name', 'Mark1', 'Mark2', 'Mark3'])
    data = [
        [1, 'John Doe', 85, 78, 92],
        [2, 'Jane Smith', 88, 90, 85],
        [3, 'Emily Davis', 75, 85, 80],
        [4, 'Michael Brown', 92, 88, 84]
    ]
    for row in data:
        ws.append(row)
    wb.save('student.xlsx')
    print("Workbook 'student.xlsx' created successfully.")

def add_total_mark():
    wb = openpyxl.load_workbook('student.xlsx')
    ws = wb.active
    ws['F1'] = 'Total_Mark'
    for row in range(2, ws.max_row + 1):
        mark1 = ws.cell(row=row, column=3).value
        mark2 = ws.cell(row=row, column=4).value
        mark3 = ws.cell(row=row, column=5).value
        total_mark = (mark1 + mark2 + mark3) * 0.6
        ws.cell(row=row, column=6, value=total_mark)
    wb.save('student.xlsx')
    print("Workbook 'student.xlsx' updated with 'Total_Mark' column.")

def remove_columns():
    wb = openpyxl.load_workbook('student.xlsx')
    ws = wb.active
    ws.delete_cols(4)
    ws.delete_cols(4)
    wb.save('student_new.xlsx')
    print("Workbook 'student_new.xlsx' created without 'Mark2' and 'Mark3' columns.")

def update_student_data():
    wb = openpyxl.load_workbook('student.xlsx')
    ws = wb.active
    student_id = int(input("Enter Student ID: "))
    new_mark1 = int(input("Enter new Mark1 value: "))
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == student_id:
            ws.cell(row=row, column=3, value=new_mark1)
            break
    wb.save('student.xlsx')
    print(f"Mark1 value for Student ID {student_id} updated successfully.")


create_workbook()
add_total_mark()
remove_columns()
update_student_data()
