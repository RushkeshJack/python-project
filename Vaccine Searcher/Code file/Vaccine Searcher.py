import requests
from openpyxl import load_workbook
from tkinter import *
from tkinter import messagebox
from tkcalendar import Calendar
import datetime
import webbrowser


def vaccine_result(): # This function will get executed when there are vaccine present in that District ID else there will error messegebox
    window5 = Tk()
    window5_width = 900
    window5_height = 500
    window5.geometry(f'{window5_width}x{window5_height}+{280}+{180}')
    window5.resizable(False, False)
    window5.title("Available Vaccine Slots")
    window5.iconbitmap(r"C:\Users\rk\Desktop\Vaccine searcher\ICON\vaccine1.ico")

    def Cowin():
        webbrowser.open("https://selfregistration.cowin.gov.in/", new=1)


    frame1 = Frame(window5, height=120, width=700, bg='#50C878', bd=1)
    frame1.place(x=0, y=0)
    frame2 = Frame(window5, height=120, width=200, bg='#00cc00',bd=1)
    frame2.place(x=700, y=0)

    state_id = s.get()
    district_id = d.get()
    date = r.get()
    mylabel = Label(frame1, text="State ID = " + str(state_id), bg='#AAFF00', font=("arial", 18), fg='black', padx=10,
                    pady=10)
    mylabel.place(x=50, y=40)
    mylabel2 = Label(frame1, text="District ID = " + str(district_id), bg='#AAFF00', font=("arial", 18), fg='black',
                     padx=10, pady=10)
    mylabel2.place(x=240, y=40)
    mylabel3 = Label(frame1, text="Date = " + str(date), bg='#AAFF00', font=("arial", 18), fg='black', padx=10, pady=10)
    mylabel3.place(x=470, y=40)

    b=Button(frame2, text="Open Co-Win site", command=Cowin, padx=5, pady=5, font=("arial", 12),bg="#4073FF")
    b.place(x=25, y=40)

    myframe = Frame(window5, relief=GROOVE, width=880, height=380, bd=1)
    myframe.place(x=0, y=120)

    canvas = Canvas(myframe)
    frame = Frame(canvas)
    myscrollbar = Scrollbar(myframe, orient="vertical", command=canvas.yview)
    canvas.configure(yscrollcommand=myscrollbar.set)

    myscrollbar.pack(side="right", fill="y")
    canvas.pack(side="left")
    canvas.create_window((0, 0), window=frame, anchor='nw')

    def myfunction(event):
        canvas.configure(scrollregion=canvas.bbox("all"), width=880, height=380)

    frame.bind("<Configure>", myfunction)

    state_id1 = s.get()
    district_id1 = d.get()

    workbook = load_workbook('state_id-district_id.xlsx')
    sheet = workbook['vaccine searcher']

    sheet.cell(row=2, column=1).value = state_id1
    sheet.cell(row=2, column=2).value = district_id1

    workbook.save('state_id-district_id.xlsx')

    x1 = 2

    Label(frame, text="Center Name", font='Helvetica 13 bold').grid(row=0, column=0, padx=0, pady=10,
                                                                                  sticky=W)
    Label(frame, text="Pincode", font='Helvetica 13 bold' ).grid(row=0, column=1, padx=0, pady=10)
    Label(frame, text="Vaccine Name", font='Helvetica 13 bold').grid(row=0, column=2, padx=0, pady=10)
    Label(frame, text="Free/Paid", font='Helvetica 13 bold').grid(row=0, column=3, padx=0, pady=10)
    Label(frame, text="Dose 1", font='Helvetica 13 bold').grid(row=0, column=4, padx=0, pady=10)
    Label(frame, text="Dose 2", font='Helvetica 13 bold').grid(row=0, column=5, padx=0, pady=10)
    dist = d.get()
    date = r.get()
    URL = 'https://cdn-api.co-vin.in/api/v2/appointment/sessions/public/findByDistrict?district_id={}&date={}'.format(
        dist, date)
    result = requests.get(URL)
    response_json = result.json()
    data = response_json["sessions"]

    for each in data:
        Hospital = each["name"]
        Label(frame, text=Hospital, font=("arial", 11)).grid(row=x1, column=0, padx=0, pady=0, sticky=W)

        pincode = str(each["pincode"])
        Label(frame, text=pincode, font=("arial", 11)).grid(row=x1, column=1, padx=20, pady=5)

        vaccine = each["vaccine"]
        Label(frame, text=vaccine, font=("arial", 11)).grid(row=x1, column=2, padx=40, pady=2)

        fee_type = str(each["fee_type"])
        Label(frame, text=fee_type, font=("arial", 11)).grid(row=x1, column=3, padx=30, pady=2)

        dose1 = str(each["available_capacity_dose1"])
        Label(frame, text=dose1, font=("arial", 11)).grid(row=x1, column=4, padx=30, pady=2)

        dose2 = str(each["available_capacity_dose2"])
        Label(frame, text=dose2, font=("arial", 11)).grid(row=x1, column=5, padx=30, pady=2)

        x1 += 1


def submit(): # This function will get executed when 'Submit' Button get clicked

    dist = d.get()
    date = r.get()
    URL = 'https://cdn-api.co-vin.in/api/v2/appointment/sessions/public/findByDistrict?district_id={}&date={}'.format(dist, date)
    result = requests.get(URL)
    response_json = result.json()
    data = response_json["sessions"]

    if(data == []):
        messagebox.showwarning("Unexpected Error","No Vaccination center is available for booking or Please cheack the State ID and District ID are correct or not!")

    elif(data != []):
        state_id1 = s.get()
        URL = 'https://cdn-api.co-vin.in/api/v2/admin/location/districts/{}'.format(state_id1)
        result = requests.get(URL)
        response_json = result.json()
        data = response_json["districts"]
        count=0
        count2=0
        for e in data:
            count+=1

        for each in data:
            district_id = each["district_id"]  # district_id is a int type but dist is string type     -> print(district_id)      -> print(dist+"z")
            count2+=1
            if(int(dist) == district_id):
                 vaccine_result()
                 break

        if(count == count2):
             messagebox.showwarning("Unexpected Error","District ID for given State ID is wrong. Please enter correct District ID!")



def restore():  # This function will get executed when 'Restore' Button get clicked

    workbook = load_workbook('state_id-district_id.xlsx')
    sheet = workbook['vaccine searcher']

    state_id = sheet['A2'].value
    district_id = sheet['B2'].value

    if (state_id == None):
        messagebox.showerror("Error","There are no data about State_id and District_id")
    else:
        s.insert(0,state_id)
        d.insert(0,district_id)



def about_us():  # This function will get executed when 'About Us' Button get clicked

    messagebox.showinfo("About Us", "We are Engineering students and working on a problem i.e. how to make human lives simpler!")



def user_manual():  # This function will get executed when 'User Manual' Button get clicked

    messagebox.showinfo("User Manual", "Step 1: Enter the State ID in box. For reference click 'State List' button\n"
                                       "Step 2: Enter the District ID in box. For reference click 'district List' button\n"
                                       "step 3: Enter date in 'DD-MM-YYYY' format or click on 'Calendar' button and Select date from Calendar")


def state_list():  # This function will get executed when 'State List' Button get clicked

  window1 = Tk()
  window1_width = 1000
  window1_height = 600
  window1.geometry(f'{window1_width}x{window1_height}+{230}+{120}')
  window1.resizable(False, False)
  window1.title("STATE LIST")
  window1.iconbitmap(r"C:\Users\rk\Desktop\Vaccine searcher\ICON\states.ico")

  frame1 = Frame(window1, height=90, width=1000, bg='#ff0080', bd=1)
  frame1.place(x=0, y=0)

  URL = 'https://cdn-api.co-vin.in/api/v2/admin/location/states'
  result = requests.get(URL)
  response_json = result.json()
  data = response_json["states"]
  y1=100
  x1=20
  x2=200
  count=0
  info1=Label(frame1,text="* Below are the ' State Name ' and in front of them ' State Id ' are given",font='Helvetica 13 bold',bg='#ff0080')
  info2=Label(frame1,text="* Please select the correct State Id and write it in ' State Id ' box",font='Helvetica 13 bold',bg='#ff0080')
  info1.place(x=20, y=20)
  info2.place(x=20, y=45)

  for each in data:

     all_state=Label(window1,text=each["state_name"],font=("arial", 10))
     all_state.place(x=x1, y=y1)
     state_id=Label(window1,text=" -  "+ str(each["state_id"]),font=("arial",10))
     state_id.place(x=x2,y=y1)
     y1+=40
     count+=1
     if(count>10):
         x1+=250
         x2+=250
         y1=100
         count=0



def district_list():  # This function will get executed when 'District List' Button get clicked

    state_id1=s.get()

    if(state_id1=="" or (int(state_id1)<0 or int(state_id1)>37)):
      window2 = Tk()
      window2_width=250
      window2_height=130
      window2.geometry(f'{window2_width}x{window2_height}+{630}+{400}')
      window2.resizable(False, False)
      window2.title("ERROR")
      window2.iconbitmap(r"C:\Users\rk\Desktop\vaccine searcher\ICON\error.ico")
      Error_messege1 = Label(window2,text="Please Select The State Id First !!")
      Error_messege2 = Label(window2,text="OR")
      Error_messege3 = Label(window2, text="State Id You Typed Is Wrong !!")
      Error_messege1.place(x=45, y=10)
      Error_messege2.place(x=110, y=50)
      Error_messege3.place(x=50, y=90)

    else:
      window3 = Tk()
      window3_width=1300
      window3_height=750
      window3.geometry(f'{window3_width}x{window3_height}+{100}+{50}')
      window3.resizable(False, False)
      window3.title("DISTRICT LIST")
      window3.iconbitmap(r"C:\Users\rk\Desktop\Vaccine searcher\ICON\districts.ico")

      frame1 = Frame(window3, height=90, width=1300, bg='#ff8000', bd=1)
      frame1.place(x=0, y=0)

      state_id1 = s.get()
      URL = 'https://cdn-api.co-vin.in/api/v2/admin/location/districts/{}'.format(state_id1)
      result = requests.get(URL)
      response_json = result.json()
      data = response_json["districts"]
      y1=100
      x1=20
      x2=200
      count=0
      info1=Label(frame1,text="* Below are the ' District Name ' and in front of them ' District Id ' are given",font='Helvetica 13 bold',bg='#ff8000')
      info2=Label(frame1,text="* Please select the correct District Id and write it in ' District Id ' box",font='Helvetica 13 bold',bg='#ff8000')
      info1.place(x=20, y=20)
      info2.place(x=20, y=45)

      for each in data:

       all_state=Label(window3,text=each["district_name"],font=("arial", 10))
       all_state.place(x=x1, y=y1)
       state_id=Label(window3,text="  -  "+ str(each["district_id"]),font=("arial",10))
       state_id.place(x=x2,y=y1)
       y1+=40
       count+=1
       if(count>15):
         x1+=250
         x2+=250
         y1=100
         count=0

def calendar_fun():  # This function will get executed when 'Calendar' Button get clicked

    window4 = Tk()
    window4.title("Calendar")
    window4_width = 300
    window4_height = 280
    window4.geometry(f'{window4_width}x{window4_height}+{600}+{300}')
    window4.resizable(False, False)
    window4.iconbitmap(r"C:\Users\rk\Desktop\Vaccine searcher\ICON\calendar.ico")
    cal = Calendar(window4, selectmode="day",year=2021,month=9,day=19,date_pattern="dd-mm-yy")
    cal.pack(padx=20,pady=20)
    def grab_date():
        r.delete(0,END)
        date = cal.get_date()
        r.insert(0,date)
        window4.destroy()
    Select_date = Button(window4, text='Select', command=grab_date, padx=1, pady=1, font=("arial", 13))
    Select_date.pack()




window = Tk()
window.title("Vaccine Searcher version 1.0")
window_width = 500
window_height = 550
window.geometry(f'{window_width}x{window_height}+{500}+{150}')
window.resizable(False, False)
window.iconbitmap(r"C:\Users\rk\Desktop\Vaccine searcher\ICON\vaccine.ico")
top_frame_bg = '#867ae9'

frame1 = Frame(window, height=80, width=500, bg=top_frame_bg, bd=1, relief=FLAT)
frame1.place(x=0, y=0)
frame2 = Frame(window, height=370, width=500, bd=1, relief=FLAT)
frame2.place(x=0, y=80)


menubar = Menu(window)
menubar.add_command(label="User Manual", command=user_manual)
menubar.add_command(label="About Us", command=about_us)
menubar.add_command(label="Quit", command=window.quit)
window.config(menu=menubar)


heading = Label(window,text="Vaccine Searcher, Search Now!",font=("arial",15),bg=top_frame_bg)
heading.place(x=110,y=25)


State = Label(window,text="State Id :",font=("arial", 15))
State.place(x=37, y=170)
s = Entry(window,width=20,font=("arial", 15))
s.place(x=140, y=175)

District = Label(window,text="District Id :",font=("arial", 15))
District.place(x=22, y=270)
d = Entry(window,width=20,font=("arial", 15))
d.place(x=140, y=275)

Date = Label(window,text="Date :",font=("arial", 15))
Date.place(x=65, y=370)
r = Entry(window,width=20,font=("arial", 15))
r.place(x=140, y=375)
x = datetime.datetime.now()
date = x.strftime("%d-%m-%y")
r.insert(0, date)

State_list = Button(window,text="State List",command=state_list,padx=1,pady=1,font=("arial", 13),).place(x=380, y=170)
District_list = Button(window,text="District List",command=district_list,padx=1,pady=1,font=("arial", 13)).place(x=380, y=270)
Date_pick = Button(window, text='Calendar', command=calendar_fun,padx=1,pady=1,font=("arial", 13)).place(x=380, y=370)

Restore = Button(window,text="Restore",command=restore,padx=5,pady=5,font=("arial", 12)).place(x=380, y=90)
Submit = Button(window,text="Submit",command=submit,padx=5,pady=5,font=("arial", 14)).place(x=220, y=430)


window.mainloop()